from openpyxl import load_workbook
import matplotlib.pyplot as plt
import numpy as np

workbook = load_workbook(filename="Stocks_1.xlsx")
print(workbook.sheetnames)
#  a
sheet = workbook.active
for i in range(2, 7):
    print(sheet.cell(row=i, column=1).value)

print(" ")

for i in range(5):
    print(sheet.cell(row=248+i, column=1).value)
#  b
print(" ")
apple = []
temp = ''
for i in range(2, 252):
    temp = sheet.cell(row=i, column=1).value
    if len(temp) == 42:
        # print(temp[16:22])
        apple.append(float(temp[16:22]))
    if len(temp) == 43:
        # print(temp[17:23])
        apple.append(float(temp[17:23]))
    if len(temp) == 44:
        # print(temp[18:24])
        apple.append(float(temp[18:24]))
#  c
print(" ")
print("length:", len(apple))
#  d
x = [i for i in range(1, 250)]
y = [apple[i] for i in range(0, len(apple)-1)]
plt.plot(x, y)
plt.grid()
plt.show()
#  e
m_apple = 0
for i in range(0, len(apple)):
    m_apple = m_apple + apple[i]
m_apple = m_apple/len(apple)
med_apple = apple[124]

print(" ")
print(m_apple, med_apple)